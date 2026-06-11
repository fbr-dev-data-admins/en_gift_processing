import streamlit as st
import pandas as pd
import json
import re
import os
import requests
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule, CellIsRule
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

from en_api import authenticate as en_authenticate
from re_skyapi import RESkyAPI
from transform import GiftTransformer

# ---------- GITHUB CONFIG LOADING ----------
def load_config_from_github(repo_url: str, file_path: str) -> dict:
    if "github.com" in repo_url:
        parts = repo_url.rstrip('/').split('/')
        owner = parts[-2]
        repo  = parts[-1].replace('.git', '')
        raw_url = f"https://raw.githubusercontent.com/{owner}/{repo}/main/{file_path}"
    else:
        raw_url = repo_url
    
    try:
        response = requests.get(raw_url, timeout=10)
        response.raise_for_status()
        return response.json()
    except requests.exceptions.RequestException as e:
        st.warning(f"Could not load {file_path} from GitHub: {e}")
        return {}
    except json.JSONDecodeError as e:
        st.warning(f"Invalid JSON in {file_path}: {e}")
        return {}

# ---------- PASSWORD PROTECTION ----------
def check_password():
    def password_entered():
        if st.session_state["password"] == st.secrets["app"]["password"]:
            st.session_state["password_correct"] = True
        else:
            st.session_state["password_correct"] = False

    if "password_correct" not in st.session_state:
        st.text_input("Password", type="password", on_change=password_entered, key="password")
        return False
    elif not st.session_state["password_correct"]:
        st.text_input("Password", type="password", on_change=password_entered, key="password")
        st.error("Password incorrect")
        return False
    else:
        return True


# ---------- HELPER FUNCTIONS ----------
def load_json_config(filepath: str) -> dict:
    """Load JSON configuration file, tolerating trailing commas (JSON5-style)"""
    if os.path.exists(filepath):
        with open(filepath, 'r') as f:
            raw = f.read()
        raw = re.sub(r',\s*([}\]])', r'\1', raw)
        return json.loads(raw)
    return {}


def load_text_list_config(filepath: str) -> list:
    """Load a plain-text config file where each line is one entry"""
    if os.path.exists(filepath):
        with open(filepath, 'r') as f:
            return [line.strip().lower() for line in f if line.strip()]
    return []


def load_text_list_from_github(repo_url: str, file_path: str) -> list:
    """Load a plain-text (one entry per line) file from a GitHub repository."""
    if "github.com" in repo_url:
        parts = repo_url.rstrip('/').split('/')
        owner = parts[-2]
        repo  = parts[-1].replace('.git', '')
        raw_url = f"https://raw.githubusercontent.com/{owner}/{repo}/main/{file_path}"
    else:
        raw_url = repo_url

    try:
        response = requests.get(raw_url, timeout=10)
        response.raise_for_status()
        return [line.strip().lower() for line in response.text.splitlines() if line.strip()]
    except requests.exceptions.RequestException as e:
        st.warning(f"Could not load {file_path} from GitHub: {e}")
        return []


def save_json_config(filepath: str, data: dict):
    """Save JSON configuration file"""
    with open(filepath, 'w') as f:
        json.dump(data, f, indent=2)


def push_config_to_github(repo_url: str, file_path: str, data: dict, token: str, commit_message: str = "Update config") -> bool:
    try:
        parts = repo_url.rstrip('/').split('/')
        owner = parts[-2]
        repo  = parts[-1].replace('.git', '')
        
        api_url = f"https://api.github.com/repos/{owner}/{repo}/contents/{file_path}"
        
        headers = {
            "Authorization": f"token {token}",
            "Accept": "application/vnd.github.v3+json"
        }
        
        response = requests.get(api_url, headers=headers)
        sha = response.json()['sha'] if response.status_code == 200 else None
        
        import base64
        content         = json.dumps(data, indent=2)
        encoded_content = base64.b64encode(content.encode()).decode()
        
        payload = {
            "message": commit_message,
            "content": encoded_content,
            "branch":  "main"
        }
        if sha:
            payload["sha"] = sha
        
        response = requests.put(api_url, headers=headers, json=payload)
        return response.status_code in [200, 201]
        
    except Exception as e:
        st.error(f"Failed to push to GitHub: {e}")
        return False


def save_config_with_github_sync(filepath: str, data: dict, github_repo: str = "", github_token: str = "", config_type: str = "config"):
    save_json_config(filepath, data)
    if github_repo and github_token:
        success = push_config_to_github(
            repo_url=github_repo,
            file_path=filepath,
            data=data,
            token=github_token,
            commit_message=f"Update {config_type} via Streamlit app"
        )
        if not success:
            st.warning(f"⚠️ Saved locally but failed to push to GitHub. Check your token permissions.")


def create_excel_output(df: pd.DataFrame, exceptions_df: pd.DataFrame = None) -> BytesIO:
    """Create Excel workbook with conditional formatting and formulas"""
    output = BytesIO()
    wb = Workbook()
    
    ws_main = wb.active
    ws_main.title = "Gift Import"
    
    headers = list(df.columns)
    
    for c_idx, header in enumerate(headers, 1):
        cell = ws_main.cell(row=1, column=c_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    col_letters = {}
    for idx, header in enumerate(headers):
        col_letters[header] = get_column_letter(idx + 1)
    
    for r_idx, (df_idx, row) in enumerate(df.iterrows(), 2):
        for c_idx, header in enumerate(headers, 1):
            value = row[header]
            
            if header == 'Addressee' and 'Last Name' in col_letters and 'Spouse Last Name' in col_letters:
                ln_col  = col_letters['Last Name']
                sln_col = col_letters['Spouse Last Name']
                formula = f'=IF({ln_col}{r_idx}={sln_col}{r_idx},49,48)'
                ws_main.cell(row=r_idx, column=c_idx, value=formula)
            
            elif header == 'Spouse Addressee' and 'Addressee' in col_letters and 'Spouse Last Name' in col_letters:
                addr_col = col_letters['Addressee']
                sln_col  = col_letters['Spouse Last Name']
                formula  = f'=IF({sln_col}{r_idx}<>"",{addr_col}{r_idx},"")'
                ws_main.cell(row=r_idx, column=c_idx, value=formula)
            
            elif header == 'Salutation' and 'Spouse Last Name' in col_letters:
                sln_col = col_letters['Spouse Last Name']
                formula = f'=IF({sln_col}{r_idx}="",35,46)'
                ws_main.cell(row=r_idx, column=c_idx, value=formula)
            
            elif header == 'Spouse Salutation' and 'Salutation' in col_letters and 'Spouse Last Name' in col_letters:
                sal_col = col_letters['Salutation']
                sln_col = col_letters['Spouse Last Name']
                formula = f'=IF({sln_col}{r_idx}<>"",{sal_col}{r_idx},"")'
                ws_main.cell(row=r_idx, column=c_idx, value=formula)
            
            elif header == 'Nickname' and 'First Name' in col_letters:
                fn_col  = col_letters['First Name']
                formula = f'=IF({fn_col}{r_idx}="","",{fn_col}{r_idx})'
                ws_main.cell(row=r_idx, column=c_idx, value=formula)
            
            elif header == 'Spouse Nickname' and 'Spouse First Name' in col_letters:
                sfn_col = col_letters['Spouse First Name']
                formula = f'=IF({sfn_col}{r_idx}<>"",{sfn_col}{r_idx},"")'
                ws_main.cell(row=r_idx, column=c_idx, value=formula)
            
            else:
                if pd.isna(value):
                    ws_main.cell(row=r_idx, column=c_idx, value='')
                else:
                    cell_obj = ws_main.cell(row=r_idx, column=c_idx, value=value)
                    if header == 'Cell':
                        cell_obj.number_format = '@'
    
    # Auto-adjust column widths
    for col in ws_main.columns:
        max_length = 0
        column     = col[0].column_letter
        for cell in col:
            try:
                cell_val = str(cell.value) if cell.value else ''
                if not cell_val.startswith('='):
                    if len(cell_val) > max_length:
                        max_length = len(cell_val)
            except:
                pass
        ws_main.column_dimensions[column].width = min(max(max_length + 2, 10), 50)
    
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    orange_fill = PatternFill(start_color="F2AC57", end_color="F2AC57", fill_type="solid")
    
    # Highlight Spouse First Name if contains number
    if "Spouse First Name" in col_letters:
        col_letter = col_letters["Spouse First Name"]
        ws_main.conditional_formatting.add(
            f'{col_letter}2:{col_letter}{len(df)+1}',
            FormulaRule(
                formula=[f'OR(ISNUMBER(SEARCH("0",{col_letter}2)),ISNUMBER(SEARCH("1",{col_letter}2)),ISNUMBER(SEARCH("2",{col_letter}2)),ISNUMBER(SEARCH("3",{col_letter}2)),ISNUMBER(SEARCH("4",{col_letter}2)),ISNUMBER(SEARCH("5",{col_letter}2)),ISNUMBER(SEARCH("6",{col_letter}2)),ISNUMBER(SEARCH("7",{col_letter}2)),ISNUMBER(SEARCH("8",{col_letter}2)),ISNUMBER(SEARCH("9",{col_letter}2)))'],
                fill=yellow_fill
            )
        )
    
    # Highlight Gifts Last Month with CHECK
    if "Gifts Last Month" in col_letters:
        col_letter = col_letters["Gifts Last Month"]
        ws_main.conditional_formatting.add(
            f'{col_letter}2:{col_letter}{len(df)+1}',
            FormulaRule(
                formula=[f'ISNUMBER(SEARCH("CHECK",{col_letter}2))'],
                fill=yellow_fill
            )
        )
    
    # Highlight RE Constituent ID and Fund ID if Key Indicator = "O"
    if "Key Indicator" in col_letters:
        ki_letter = col_letters["Key Indicator"]
        
        if "Constituent ID" in col_letters:
            col_letter = col_letters["Constituent ID"]
            ws_main.conditional_formatting.add(
                f'{col_letter}2:{col_letter}{len(df)+1}',
                FormulaRule(formula=[f'${ki_letter}2="O"'], fill=orange_fill)
            )
        
        if "Fund ID" in col_letters:
            col_letter = col_letters["Fund ID"]
            ws_main.conditional_formatting.add(
                f'{col_letter}2:{col_letter}{len(df)+1}',
                FormulaRule(formula=[f'${ki_letter}2="O"'], fill=orange_fill)
            )
    
    # Highlight Constituent ID if contains "%" (custom note appended)
    if "Constituent ID" in col_letters:
        custom_note_fill = PatternFill(start_color="92CDDC", end_color="92CDDC", fill_type="solid")
        col_letter = col_letters["Constituent ID"]
        ws_main.conditional_formatting.add(
            f'{col_letter}2:{col_letter}{len(df)+1}',
            FormulaRule(
                formula=[f'AND({col_letter}2<>"",ISNUMBER(SEARCH("%",{col_letter}2)))'],
                fill=custom_note_fill
            )
        )
    
    # Highlight blank Country when any address field has data
    if "Country" in col_letters:
        country_col = col_letters["Country"]
        addr_col    = col_letters.get("Address", "")
        city_col    = col_letters.get("City", "")
        state_col   = col_letters.get("State", "")
        zip_col     = col_letters.get("ZIP", "")
        
        conditions = []
        if addr_col:  conditions.append(f'{addr_col}2<>""')
        if city_col:  conditions.append(f'{city_col}2<>""')
        if state_col: conditions.append(f'{state_col}2<>""')
        if zip_col:   conditions.append(f'{zip_col}2<>""')
        
        if conditions:
            or_formula = ",".join(conditions)
            formula    = f'AND({country_col}2="",OR({or_formula}))'
            ws_main.conditional_formatting.add(
                f'{country_col}2:{country_col}{len(df)+1}',
                FormulaRule(formula=[formula], fill=yellow_fill)
            )
    
    # Highlight blank City when Address AND ZIP both have data
    if "City" in col_letters:
        city_col = col_letters["City"]
        addr_col = col_letters.get("Address", "")
        zip_col  = col_letters.get("ZIP", "")
        
        if addr_col and zip_col:
            formula = f'AND({city_col}2="",{addr_col}2<>"",{zip_col}2<>"")'
            ws_main.conditional_formatting.add(
                f'{city_col}2:{city_col}{len(df)+1}',
                FormulaRule(formula=[formula], fill=yellow_fill)
            )
    
    # Add exceptions sheet
    if exceptions_df is not None and len(exceptions_df) > 0:
        ws_exceptions = wb.create_sheet("Exceptions")
        for r_idx, row in enumerate(dataframe_to_rows(exceptions_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_exceptions.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    # Highlight entire row for DAF gifts (Gift Subtype = Chariot)
    daf_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    if 'Gift Subtype' in col_letters:
        gs_col         = col_letters['Gift Subtype']
        last_col_letter = get_column_letter(len(headers))
        ws_main.conditional_formatting.add(
            f'A2:{last_col_letter}{len(df)+1}',
            FormulaRule(
                formula=[f'${gs_col}2="Chariot"'],
                fill=daf_fill
            )
        )
    
    wb.save(output)
    output.seek(0)
    return output


def create_final_import_file(df: pd.DataFrame) -> BytesIO:
    """Create final import file with specified column order"""
    final_columns = [
        "Constituent ID", "Engaging Networks ID", "First Name", "Middle Name", "Last Name",
        "Spouse First Name", "Spouse Middle Name", "Spouse Last Name", "Addressee", "Salutation",
        "Spouse Addressee", "Spouse Salutation", "Address", "City", "State", "ZIP", "Country",
        "E-mail", "Cell", "EN Transaction ID", "Gift Date", "Gift Amount",
        "Gift Reference", "Gift Solicitor", "Campaign", "Appeal ID", "Fund ID", "Package", "Branch",
        "Gift Subtype", "Credit Type", "Donation Type", "Stripe Transaction ID", "EN Donation Form Name",
        "EN Campaign ID", "EN Campaign Name", "EN Fundraising Page ID", "EN Fundraising Page Name",
        "Monthly Donor Status Description", "Monthly Donor Status Date", "Monthly Donor Anniversary Description",
        "Monthly Donor Anniversary Date", "Monthly Donor Annual Statement Type", "Monthly Donor Channel",
        "Monthly Donor Payment Method", "Monthly Donor Region",
        "Consents Date", "Email Channel Opt-In", "Email Direct Marketing Consent",
        "Email Direct Marketing Consent Statement", "Email Fundraising Consent",
        "Email Fundraising Consent Statement", "Email Newsletter Consent",
        "Email Newsletter Consent Statement", "SMS Channel Response", "SMS Consent Statement"
    ]
    
    for col in final_columns:
        if col not in df.columns:
            df[col] = ""
    
    available_cols = [col for col in final_columns if col in df.columns]
    df_final = df[available_cols].copy()
    
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Import"
    
    for r_idx, row in enumerate(dataframe_to_rows(df_final, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.font = Font(bold=True)
    
    wb.save(output)
    output.seek(0)
    return output


# ---------- MAIN APP ----------
if check_password():
    st.set_page_config(page_title="EN Gift Transformation", layout="wide")
    
    st.title("🎁 Engaging Networks Gift Transformation")
    
    # Initialize session state
    if 'transformer' not in st.session_state:
        st.session_state.transformer = GiftTransformer()
    if 'processed_df' not in st.session_state:
        st.session_state.processed_df = None
    if 'exceptions_df' not in st.session_state:
        st.session_state.exceptions_df = None
    if 'p2p_pending' not in st.session_state:
        st.session_state.p2p_pending = []
    if 're_api' not in st.session_state:
        try:
            st.session_state.re_api = RESkyAPI(
                client_id=st.secrets["re_api"]["client_id"],
                client_secret=st.secrets["re_api"]["client_secret"],
                redirect_uri=st.secrets["re_api"]["redirect_uri"],
                subscription_key=st.secrets["re_api"]["subscription_key"]
            )
        except Exception as e:
            st.session_state.re_api = None
    
    mapping_path      = "config/mapping.json"
    p2p_path          = "config/P2P.json"
    custom_notes_path = "config/custom_notes.json"
    
    github_repo  = st.secrets.get("github", {}).get("config_repo", "")
    github_token = st.secrets.get("github", {}).get("access_token", "")
    
    md_acks_path = "config/md_acks.txt"

    if github_repo:
        if 'mapping_config' not in st.session_state or st.sidebar.button("🔄 Reload GitHub Configs"):
            st.session_state.mapping_config      = load_config_from_github(github_repo, "config/mapping.json")
            st.session_state.p2p_config          = load_config_from_github(github_repo, "config/P2P.json")
            st.session_state.custom_notes_config = load_config_from_github(github_repo, "config/custom_notes.json")
            st.session_state.md_acks_config      = load_text_list_from_github(github_repo, "config/md_acks.txt")
        mapping_config      = st.session_state.get('mapping_config', {})
        p2p_config          = st.session_state.get('p2p_config', {})
        custom_notes_config = st.session_state.get('custom_notes_config', {})
        md_acks_config      = st.session_state.get('md_acks_config', [])
    else:
        mapping_config      = load_json_config(mapping_path)
        p2p_config          = load_json_config(p2p_path)
        custom_notes_config = load_json_config(custom_notes_path)
        md_acks_config      = load_text_list_config(md_acks_path)
    
    # Sidebar
    with st.sidebar:
        st.header("⚙️ Configuration")
        
        if github_repo:
            st.success(f"📂 Configs from GitHub")
            st.caption(github_repo)
        else:
            st.info("📂 Configs from local files")
        
        st.divider()
        
        st.subheader("RE Sky API Status")
        if st.session_state.re_api:
            if st.session_state.re_api.is_authenticated():
                st.success("✅ Authenticated")
            else:
                st.warning("⚠️ Not authenticated")
                
                if st.button("🔑 Get Authorization URL"):
                    auth_url = st.session_state.re_api.get_authorization_url()
                    st.session_state.re_auth_url = auth_url
                
                if 're_auth_url' in st.session_state:
                    st.markdown(f"**Step 1:** [Click here to authorize]({st.session_state.re_auth_url})")
                    st.markdown("**Step 2:** After authorizing, copy the code from the URL")
                    
                    auth_code = st.text_input(
                        "Authorization Code:",
                        key="re_auth_code",
                        placeholder="Paste the code here"
                    )
                    st.caption("Press the button below after pasting the code")
                    
                    if st.button("✅ Submit Authorization Code"):
                        if auth_code:
                            with st.spinner("Exchanging code for token..."):
                                if st.session_state.re_api.exchange_code_for_token(auth_code):
                                    st.success("🎉 Authentication successful!")
                                    del st.session_state.re_auth_url
                                    st.rerun()
                                else:
                                    st.error("❌ Failed to authenticate. Check the code and try again.")
                        else:
                            st.error("Please enter the authorization code first.")
        else:
            st.error("❌ RE API not configured")
            st.caption("Add RE API credentials to secrets.toml")
    
    # Main content tabs
    tab1, tab2, tab3 = st.tabs(["📥 Data Export", "🔄 Transform", "👥 P2P Matching"])
    
    # ---------- TAB 1: DATA EXPORT ----------
    with tab1:
        st.header("Step 1: Export Data from Engaging Networks")
        
        yesterday = datetime.today() - timedelta(days=1)
        
        col1, col2 = st.columns(2)
        with col1:
            start_date = st.date_input("Start Date", yesterday)
        with col2:
            end_date = st.date_input("End Date", yesterday)
        
        if st.button("🔍 Fetch EN Data", type="primary"):
            try:
                token = st.secrets["en_api"]["token"]
            except KeyError:
                st.error("❌ EN API token not found in secrets. Please add [en_api] section with 'token' to your secrets.toml")
                token = None
            
            if token:
                with st.spinner("Fetching data from Engaging Networks..."):
                    try:
                        start_str = start_date.strftime("%m%d%Y")
                        end_str   = end_date.strftime("%m%d%Y")
                        
                        rows = en_authenticate(token, start_str, end_str)
                        
                        if isinstance(rows, pd.DataFrame):
                            df = rows
                        else:
                            df = pd.DataFrame(rows[1:], columns=rows[0])
                        
                        st.session_state.raw_df        = df
                        st.session_state.en_start_date = start_date
                        st.session_state.en_end_date   = end_date
                        st.success(f"✅ Retrieved {len(df)} records!")
                        st.dataframe(df.head(20))
                        
                    except Exception as e:
                        st.error(f"Error fetching data: {e}")
                        st.info("💡 Make sure your EN API token is correct and has bulk export permissions.")
    
    # ---------- TAB 2: TRANSFORM ----------
    with tab2:
        st.header("Step 2: Transform Data")
        
        if 'raw_df' not in st.session_state or st.session_state.raw_df is None:
            st.warning("⚠️ Please fetch or upload data in Step 1 first.")
        else:
            df = st.session_state.raw_df.copy()
            
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Total Records", len(df))
            with col2:
                if 'Campaign Type' in df.columns:
                    st.metric("Campaign Types", df['Campaign Type'].nunique())
            with col3:
                if 'Campaign Status' in df.columns:
                    st.metric("Successful Transactions", len(df[df['Campaign Status'] == 'success']))
            
            # Check for missing form names in mapping
            if 'Campaign ID' in df.columns and 'Campaign Type' in df.columns and mapping_config:
                valid_success_types = ['FCS', 'FBS', 'FCR', 'FBR', 'PFCS', 'PFBS', 'PFCR', 'PFBR']
                valid_pending_types = ['FBS', 'FBR', 'PFBS', 'PFBR']
                
                if 'Campaign Status' in df.columns:
                    processable_mask = (
                        (df['Campaign Type'].isin(valid_success_types) & (df['Campaign Status'] == 'success')) |
                        (df['Campaign Type'].isin(valid_pending_types) & (df['Campaign Status'] == 'pending'))
                    )
                else:
                    processable_mask = df['Campaign Type'].isin(valid_success_types)
                
                processable_df = df[processable_mask]
                excluded_forms_mask = processable_df['Campaign ID'].str.startswith(('D.8.', 'Y.8.', 'S.8.'), na=False)
                processable_df = processable_df[~excluded_forms_mask]
                
                form_names     = processable_df['Campaign ID'].unique()
                fy_designation = mapping_config.get('fiscal_year_designation', '')
                forms          = mapping_config.get('forms', {})
                missing_forms  = [f for f in form_names if f not in forms and pd.notna(f)]
                
                if missing_forms:
                    st.warning(f"⚠️ Missing form names in mapping.json: {missing_forms}")
                    
                    with st.expander("Add Missing Form Mappings"):
                        for form_name in missing_forms[:10]:
                            st.subheader(f"Form: {form_name}")
                            col1, col2, col3 = st.columns(3)
                            with col1:
                                appeal = st.text_input(f"Appeal for {form_name} (do not include FY number)", key=f"appeal_{form_name}")
                            with col2:
                                fund = st.text_input(f"Fund for {form_name}", key=f"fund_{form_name}")
                            with col3:
                                is_match = st.checkbox(f"MATCH package?", key=f"match_{form_name}")
                            
                            if st.button(f"Save {form_name}", key=f"save_{form_name}"):
                                if 'forms' not in mapping_config:
                                    mapping_config['forms'] = {}
                                
                                mapping_config['forms'][form_name] = {'appeal': appeal, 'fund': fund}
                                
                                if is_match:
                                    if 'MATCH' not in mapping_config:
                                        mapping_config['MATCH'] = []
                                    if form_name not in mapping_config['MATCH']:
                                        mapping_config['MATCH'].append(form_name)
                                
                                save_config_with_github_sync(mapping_path, mapping_config, github_repo, github_token, "mapping config")
                                st.session_state.mapping_config = mapping_config
                                st.success(f"✅ Saved mapping for {form_name}" + (" (with MATCH)" if is_match else ""))
                                st.rerun()
            
            if st.button("🔄 Run Transformation", type="primary"):
                with st.spinner("Transforming data..."):
                    try:
                        transformer = st.session_state.transformer
                        
                        tribute_df = None
                        if 'raw_df' in st.session_state:
                            tribute_df = st.session_state.raw_df[
                                st.session_state.raw_df['Campaign Type'].isin(['FIM', 'PFIM'])
                            ].copy() if 'Campaign Type' in st.session_state.raw_df.columns else None
                        
                        cached_gifts      = {}
                        cached_gifts_debug = {}
                        batch_fetch_debug = {
                            're_api_available':           bool(st.session_state.re_api),
                            're_api_authenticated':       False,
                            'has_recurring_records':      False,
                            'recurring_count':            0,
                            'campaign_date_column_exists': 'Campaign Date' in df.columns,
                            'valid_dates_count':           0,
                            'date_range_calculated':       False,
                            'cache_key':                   None,
                            'used_cached':                 False,
                            'api_call_made':               False,
                            'api_call_result':             None,
                            'error':                       None
                        }
                        
                        re_api = st.session_state.re_api
                        if re_api:
                            batch_fetch_debug['re_api_authenticated'] = re_api.is_authenticated()
                        
                        if re_api and re_api.is_authenticated():
                            has_recurring = False
                            if 'Campaign Type' in df.columns:
                                recurring_mask = df['Campaign Type'].isin(['FCR', 'FBR', 'PFCR', 'PFBR'])
                                has_recurring  = recurring_mask.any()
                                batch_fetch_debug['has_recurring_records'] = has_recurring
                                batch_fetch_debug['recurring_count']       = int(recurring_mask.sum())
                            
                            if has_recurring and 'Campaign Date' in df.columns:
                                campaign_dates = pd.to_datetime(df['Campaign Date'], errors='coerce')
                                valid_dates    = campaign_dates.dropna()
                                batch_fetch_debug['valid_dates_count'] = len(valid_dates)
                                
                                if len(valid_dates) > 0:
                                    min_date = valid_dates.min()
                                    max_date = valid_dates.max()
                                    
                                    prev_min_date = min_date - relativedelta(months=1)
                                    prev_max_date = max_date - relativedelta(months=1)
                                    
                                    start_date_str = prev_min_date.strftime('%Y-%m-%d')
                                    end_date_str   = prev_max_date.strftime('%Y-%m-%d')
                                    
                                    batch_fetch_debug['date_range_calculated']  = True
                                    batch_fetch_debug['en_date_range']          = f"{min_date.strftime('%Y-%m-%d')} to {max_date.strftime('%Y-%m-%d')}"
                                    batch_fetch_debug['prev_month_date_range']  = f"{start_date_str} to {end_date_str}"
                                    
                                    cache_key = f"{start_date_str}_{end_date_str}"
                                    batch_fetch_debug['cache_key'] = cache_key
                                    
                                    if 'gifts_cache' not in st.session_state:
                                        st.session_state.gifts_cache = {}
                                    if 'gifts_cache_debug' not in st.session_state:
                                        st.session_state.gifts_cache_debug = {}
                                    
                                    if cache_key in st.session_state.gifts_cache and len(st.session_state.gifts_cache[cache_key]) > 0:
                                        cached_gifts = st.session_state.gifts_cache[cache_key]
                                        batch_fetch_debug['used_cached']              = True
                                        batch_fetch_debug['cached_constituents_count'] = len(cached_gifts)
                                        if cache_key in st.session_state.gifts_cache_debug:
                                            batch_fetch_debug['original_api_call_result'] = st.session_state.gifts_cache_debug[cache_key]
                                    else:
                                        batch_fetch_debug['api_call_made'] = True
                                        try:
                                            cached_gifts, cached_gifts_debug = re_api.get_gifts_for_date_range(
                                                start_date=start_date_str,
                                                end_date=end_date_str
                                            )
                                            batch_fetch_debug['api_call_result']      = cached_gifts_debug
                                            batch_fetch_debug['gifts_fetched']        = cached_gifts_debug.get('total_gifts_fetched', 0)
                                            batch_fetch_debug['unique_constituents']  = len(cached_gifts)
                                            
                                            if len(cached_gifts) > 0:
                                                st.session_state.gifts_cache[cache_key]       = cached_gifts
                                                st.session_state.gifts_cache_debug[cache_key] = cached_gifts_debug
                                            else:
                                                batch_fetch_debug['cache_note'] = 'Empty result NOT cached - will retry on next run'
                                        except Exception as e:
                                            import traceback
                                            batch_fetch_debug['error']           = str(e)
                                            batch_fetch_debug['error_type']      = type(e).__name__
                                            batch_fetch_debug['error_traceback'] = traceback.format_exc()
                        
                        st.session_state.batch_fetch_debug = batch_fetch_debug
                        
                        processed_df, exceptions_df, p2p_pending = transformer.transform(
                            df=df,
                            mapping_config=mapping_config,
                            p2p_config=p2p_config,
                            tribute_df=tribute_df,
                            re_api=st.session_state.re_api,
                            cached_gifts=cached_gifts,
                            custom_notes_config=custom_notes_config,
                            md_acks=md_acks_config
                        )
                        
                        if transformer.p2p_config_updates:
                            save_config_with_github_sync(p2p_path, p2p_config, github_repo, github_token, "P2P config")
                            st.session_state.p2p_config = p2p_config
                            st.info(f"✅ Auto-matched {len(transformer.p2p_config_updates)} P2P solicitor(s) and updated config")
                        
                        st.session_state.processed_df  = processed_df
                        st.session_state.exceptions_df = exceptions_df
                        st.session_state.p2p_pending   = p2p_pending
                        
                        st.success(f"✅ Transformation complete!")
                        
                        col1, col2, col3 = st.columns(3)
                        with col1:
                            st.metric("Processed Records", len(processed_df))
                        with col2:
                            st.metric("Exceptions", len(exceptions_df) if exceptions_df is not None else 0)
                        with col3:
                            st.metric("P2P Pending Match", len(p2p_pending))
                        
                        st.dataframe(processed_df, height=400)
                        
                        if exceptions_df is not None and len(exceptions_df) > 0:
                            with st.expander("View Exceptions"):
                                st.dataframe(exceptions_df)
                    
                    except Exception as e:
                        st.error(f"Transformation error: {e}")
                        import traceback
                        st.code(traceback.format_exc())
            
            st.divider()
            st.subheader("📤 Export Working File")
            
            if st.session_state.processed_df is not None and len(st.session_state.processed_df) > 0:
                if st.session_state.p2p_pending and len(st.session_state.p2p_pending) > 0:
                    st.error(f"⛔ **Cannot generate working file:** {len(st.session_state.p2p_pending)} P2P records need matching first. Please complete P2P matching in Tab 3 before exporting.")
                else:
                    st.warning(r"⚠️ **Important:** Download the working file below, then use VBA macro stored at L:\Development Systems\Database\Engaging Networks\Custom EN Transactions into RE\{Version #.#} Import Preparation.xlsm to complete final processing before RE import.")
                    
                    if st.button("📊 Generate Working Excel", type="primary"):
                        excel_buffer = create_excel_output(
                            st.session_state.processed_df,
                            st.session_state.exceptions_df
                        )
                        st.download_button(
                            label="📥 Download Working File",
                            data=excel_buffer,
                            file_name=f"EN_Gift_Transform_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
            else:
                st.info("Run transformation first to enable export.")
    
    # ---------- TAB 3: P2P MATCHING ----------
    with tab3:
        st.header("Step 3: P2P Solicitor Matching")
        
        if not st.session_state.p2p_pending:
            st.success("✅ No P2P records pending matching.")
        else:
            st.warning(f"⚠️ {len(st.session_state.p2p_pending)} records need P2P matching")
            
            for i, record in enumerate(st.session_state.p2p_pending):
                campaign_type = record.get('campaign_type', '')
                campaign_num  = record.get('campaign_number', '')
                
                if campaign_type == 'PFTC':
                    with st.expander(f"PFTC Record - Row {record.get('row_number', 'N/A')}: {record.get('campaign_data_10', 'Unknown')}"):
                        col1, col2 = st.columns(2)
                        
                        with col1:
                            st.subheader("EN Data")
                            st.write(f"**Row Number:** {record.get('row_number', '')}")
                            st.write(f"**Campaign Number:** {campaign_num}")
                            st.write(f"**Campaign Data 6:** {record.get('campaign_data_6', '')}")
                            st.write(f"**Campaign Data 7:** {record.get('campaign_data_7', '')}")
                            st.write(f"**Campaign Data 10 (Name):** {record.get('campaign_data_10', '')}")
                            st.write(f"**Campaign Data 11 (Email):** {record.get('campaign_data_11', '')}")
                        
                        with col2:
                            st.subheader("RE Match")
                            if record.get('re_match'):
                                match = record['re_match']
                                st.write(f"**RE Constituent ID:** {match.get('id', '')}")
                                st.write(f"**Name:** {match.get('name', '')}")
                                st.write(f"**Matched on:** {match.get('matched_on', '')}")
                                
                                if st.button(f"✅ Accept Match", key=f"accept_{campaign_num}"):
                                    p2p_config[campaign_num] = {
                                        'EN Campaign Name': record.get('campaign_data_10', ''),
                                        'Solicitor': match['id']
                                    }
                                    save_config_with_github_sync(p2p_path, p2p_config, github_repo, github_token, "P2P config")
                                    st.session_state.p2p_config = p2p_config
                                    st.session_state.p2p_pending.pop(i)
                                    st.success("Match saved!")
                                    st.rerun()
                            else:
                                st.warning("No automatic match found")
                            
                            manual_id = st.text_input(f"Enter RE Constituent ID manually:", key=f"manual_id_{campaign_num}")
                            if st.button(f"💾 Save Manual Match", key=f"save_manual_{campaign_num}"):
                                if manual_id:
                                    p2p_config[campaign_num] = {
                                        'EN Campaign Name': record.get('campaign_data_10', ''),
                                        'Solicitor': manual_id
                                    }
                                    save_config_with_github_sync(p2p_path, p2p_config, github_repo, github_token, "P2P config")
                                    st.session_state.p2p_config = p2p_config
                                    st.session_state.p2p_pending.pop(i)
                                    st.success("Manual match saved!")
                                    st.rerun()
                else:
                    with st.expander(f"{campaign_type} Record - Row {record.get('row_number', 'N/A')}: Campaign {campaign_num}"):
                        st.subheader("P2P Gift Details")
                        st.write(f"**Row Number:** {record.get('row_number', '')}")
                        st.write(f"**Campaign Number:** {campaign_num}")
                        st.write(f"**Campaign Type:** {campaign_type}")
                        st.write(f"**Region Page:** {record.get('campaign_id', '')}")
                        
                        st.info("""
**Please look up additional details in the P2P module for this region:**

Peer-to-Peer → Sites → Edit identified site (square and pencil) → Teams → Match on Page ID → Input Page Name and the identified solicitor RE Constituent ID

or if not found there:

Peer-to-Peer → Sites → Edit identified site (square and pencil) → Fundraisers → Match on Page ID → Input Page Name and the identified solicitor RE Constituent ID
                        """)
                        
                        st.subheader("Manual Entry Required")
                        manual_campaign_name = st.text_input(f"EN Campaign Name:", key=f"campaign_name_{campaign_num}")
                        manual_id            = st.text_input(f"RE Constituent ID (Solicitor):", key=f"manual_id_{campaign_num}")
                        
                        if st.button(f"💾 Save P2P Match", key=f"save_manual_{campaign_num}"):
                            if manual_id and manual_campaign_name:
                                p2p_config[campaign_num] = {
                                    'EN Campaign Name': manual_campaign_name,
                                    'Solicitor': manual_id
                                }
                                save_config_with_github_sync(p2p_path, p2p_config, github_repo, github_token, "P2P config")
                                st.session_state.p2p_config = p2p_config
                                st.session_state.p2p_pending.pop(i)
                                st.success("P2P match saved!")
                                st.rerun()
                            else:
                                st.error("Please enter both EN Campaign Name and RE Constituent ID")
